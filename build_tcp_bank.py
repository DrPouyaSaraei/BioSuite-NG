"""
build_tcp_bank.py
Regenerates data/tcp_parameter_bank.json from the bundled evidence
spreadsheet data/Radiobiological_TCP_NTCP.xlsx (sheet "TCP_Parameter_Bank").

--- Replaces earlier parameter-bank source (changelog) ---
This script previously read a workbook with separate TCP_Parameters /
TCP_Sources / TCP_Derived_Values sheets. That workbook has been replaced
by data/Radiobiological_TCP_NTCP.xlsx, which unifies everything (alpha,
alpha/beta, clonogen density/K, repopulation & repair parameters,
citation, verification notes) into one flat, per-record sheet. This
script has been rewritten end-to-end for that new schema; the output JSON
schema (core.tcp_bank.TCPBankEntry) is unchanged, so nothing elsewhere in
the app needed to change.

--- Model-family / computability classification ---
BioSuite-NG's "Add from TCP parameter bank" dialog always imports a
record as a standard Marsden (LQ-Poisson + linear-delay accelerated
repopulation) endpoint -- it does NOT implement every formalism a source
might report. So the important question per record isn't "what does the
free-text Model Family column say" (phrasing varies a lot row to row --
regex-matching it directly turned out to under-classify many rows in
this new workbook), it's: does this record supply data BioSuite-NG's
engine can actually use, and is its underlying formalism even compatible
with being force-fit into a Marsden model at all?

This script therefore classifies DATA-DRIVEN, not text-pattern-driven:
  1. alpha or alpha/beta missing entirely -> "incomplete_reference_only",
     not selectable, regardless of what the Model Family text says.
  2. Model Family mentions Zaider-Minerbo (a genuinely different,
     time-dependent re-sensitisation formalism BioSuite-NG's engine
     cannot reproduce) -> "incompatible_reference_only", not selectable.
  3. A source-specific clonogen density or total-K is reported ->
     "computable_full", selectable ("lq_protraction_repair" if the
     source also reports sublethal-repair/dose-protraction detail, else
     "poisson_lq_repopulation").
  4. Otherwise (alpha & alpha/beta present, no density/K reported) ->
     "alpha_beta_evidence_only", selectable -- the user supplies density,
     spread and repopulation explicitly (exactly what that status is for).
Status-key vocabulary is unchanged from the previous bank, so
ui/dialog_tcp_bank_docs.py's existing "Model-family gating" explanation
and ui/dialog_add_tcp_from_bank.py's family_note dict still apply as-is.

--- The K-vs-density safety rule (unchanged, still load-bearing) ---
A prior audit found a record (Wang JZ, Li XA 2003, prostate) whose own
text says "K=3.0e6 is total clonogens ... not a density; do not divide
by volume without a source-defined volume" -- yet the bank-import dialog
would still divide K by an arbitrary user-typed GTV volume, silently
producing a wrong (near-zero) TCP. k_fixed_no_reference_volume=True
disables that division (see ui/dialog_add_tcp_from_bank.py); the same
caveat-text regex that caught this before is reused here unchanged.

Run directly to rebuild the bank:
    python build_tcp_bank.py
"""
import json
import re
import sys
from pathlib import Path

import openpyxl

HERE = Path(__file__).resolve().parent
XLSX_PATH = HERE / "data" / "Radiobiological_TCP_NTCP.xlsx"
SHEET_NAME = "TCP_Parameter_Bank"
OUT_PATH = HERE / "data" / "tcp_parameter_bank.json"

COLUMNS = [
    "record_id", "tumour_site", "histology_setting", "modality_schedule",
    "model_family_raw", "endpoint", "alpha_definition", "alpha_per_gy",
    "alpha_value_status", "alpha_spread_sd", "alpha_beta_gy",
    "clonogen_density_percc", "total_clonogens_k", "repopulation_as_reported",
    "repopulation_days_tpot", "repopulation_rate_per_day",
    "delay_before_repopulation_days", "repair_halftime_min",
    "sublethal_repair_mu_per_min", "fraction_duration_correction",
    "source_id", "authors", "title", "journal", "doi_link", "notes",
]

K_FIXED_CAVEAT_RE = re.compile(
    r"not a density|do not divide by volume|no source-defined volume", re.IGNORECASE
)
PROTRACTION_RE = re.compile(
    r"protraction|sublethal|repair.{0,15}proliferation", re.IGNORECASE
)
ZAIDER_MINERBO_RE = re.compile(r"zaider|minerbo", re.IGNORECASE)


def parse_float(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text or text.casefold() in ("not reported", "nan", "n/a", "-"):
        return None
    try:
        return float(text)
    except ValueError:
        m = re.match(r"^-?\d+(\.\d+)?", text)
        return float(m.group(0)) if m else None


def first_author_label(authors_text: str) -> str:
    authors_text = (authors_text or "").strip()
    if not authors_text:
        return ""
    first = authors_text.split(",")[0].strip()
    first = re.sub(r"\s+et\s+al\.?$", "", first, flags=re.IGNORECASE).strip()
    if "et al" in authors_text.casefold() or "," in authors_text or "&" in authors_text:
        return f"{first} et al."
    return first


def build_citation(authors, title, journal) -> str:
    parts = []
    authors = (str(authors).strip() if authors else "")
    title = (str(title).strip() if title else "")
    journal = (str(journal).strip() if journal and str(journal).casefold() != "nan" else "")
    if authors:
        parts.append(authors.rstrip(".") + ".")
    if title:
        parts.append(title.rstrip(".") + ".")
    if journal:
        parts.append(journal.rstrip(".") + ".")
    return " ".join(parts).strip()


def classify(model_family_raw: str, alpha, alpha_beta, density, total_k):
    """Returns (model_family_key, status, selectable) -- see module
    docstring for the data-driven reasoning."""
    # Formalism-incompatible families are tagged as such regardless of data
    # completeness (a Zaider-Minerbo record missing alpha/beta is still a
    # Zaider-Minerbo record -- the label should say so, not "unclassified").
    if ZAIDER_MINERBO_RE.search(model_family_raw or ""):
        return "zaider_minerbo", "incompatible_reference_only", False
    if alpha is None or alpha_beta is None:
        return "unclassified", "incomplete_reference_only", False
    if density is not None or total_k is not None:
        if PROTRACTION_RE.search(model_family_raw or ""):
            return "lq_protraction_repair", "computable_full", True
        return "poisson_lq_repopulation", "computable_full", True
    return "alpha_beta_evidence_only", "alpha_beta_evidence_only", True


def main():
    if not XLSX_PATH.exists():
        sys.exit(f"Evidence spreadsheet not found: {XLSX_PATH}")

    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        sys.exit(f"Sheet '{SHEET_NAME}' not found in {XLSX_PATH}. Sheets present: {wb.sheetnames}")
    ws = wb[SHEET_NAME]

    rows = list(ws.iter_rows(values_only=True))
    data_rows = rows[1:]

    records = []
    family_counts = {}

    for raw in data_rows:
        if raw is None or all(v is None for v in raw):
            continue
        rec = dict(zip(COLUMNS, raw))

        alpha = parse_float(rec["alpha_per_gy"])
        alpha_beta = parse_float(rec["alpha_beta_gy"])
        density = parse_float(rec["clonogen_density_percc"])
        total_k = parse_float(rec["total_clonogens_k"])
        notes_text = str(rec["notes"] or "").strip()

        family_key, status, selectable = classify(
            rec["model_family_raw"], alpha, alpha_beta, density, total_k
        )
        family_counts[family_key] = family_counts.get(family_key, 0) + 1

        entry = {
            "record_id": str(rec["record_id"] or "").strip(),
            "tumour_site": str(rec["tumour_site"] or "").strip(),
            "histology_setting": str(rec["histology_setting"] or "").strip(),
            "modality_schedule": str(rec["modality_schedule"] or "").strip(),
            "model_family_raw": str(rec["model_family_raw"] or "").strip(),
            "model_family_key": family_key,
            "endpoint": str(rec["endpoint"] or "").strip(),
            "label": first_author_label(rec["authors"]),
            "alpha_definition": str(rec["alpha_definition"] or "").strip(),
            "alpha_per_gy": alpha,
            "alpha_value_status": str(rec["alpha_value_status"] or "").strip(),
            "alpha_spread_sd": parse_float(rec["alpha_spread_sd"]),
            "alpha_beta_gy": alpha_beta,
            "clonogen_density_percc": density,
            "total_clonogens_k": total_k,
            "k_fixed_no_reference_volume": bool(K_FIXED_CAVEAT_RE.search(notes_text)),
            "repopulation_as_reported": str(rec["repopulation_as_reported"] or "").strip(),
            "repopulation_days_tpot": parse_float(rec["repopulation_days_tpot"]),
            "repopulation_rate_per_day": parse_float(rec["repopulation_rate_per_day"]),
            "derived_mapping_basis": "",
            "delay_before_repopulation_days": parse_float(rec["delay_before_repopulation_days"]),
            "repair_halftime_min": parse_float(rec["repair_halftime_min"]),
            "sublethal_repair_mu_per_min": parse_float(rec["sublethal_repair_mu_per_min"]),
            "fraction_duration_correction": str(rec["fraction_duration_correction"] or "").strip(),
            "status": status,
            "selectable": selectable,
            "source": build_citation(rec["authors"], rec["title"], rec["journal"]),
            "url": (str(rec["doi_link"]).strip() if rec["doi_link"] and str(rec["doi_link"]).casefold() != "nan" else None),
            "notes": notes_text,
        }
        records.append(entry)

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(OUT_PATH, "w", encoding="utf-8") as f:
        json.dump(records, f, indent=2, ensure_ascii=False)

    print(f"Wrote {len(records)} TCP bank records to {OUT_PATH}")
    print(f"  selectable: {sum(1 for r in records if r['selectable'])}, "
          f"not selectable: {sum(1 for r in records if not r['selectable'])}")
    print(f"  model_family_key counts: {family_counts}")
    k_fixed = [r["record_id"] for r in records if r["k_fixed_no_reference_volume"]]
    if k_fixed:
        print(f"  k_fixed_no_reference_volume=True for: {', '.join(k_fixed)}")


if __name__ == "__main__":
    main()
