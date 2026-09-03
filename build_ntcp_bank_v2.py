"""
build_ntcp_bank_v2.py
Regenerates data/lkb_parameter_bank.json from the bundled evidence
spreadsheet data/Radiobiological_TCP_NTCP.xlsx (sheet "NTCP_Parameter_Bank").

--- Replaces earlier parameter-bank source (changelog) ---
This script previously read a workbook with separate NTCP_LKB_Parameters /
NTCP_Sources / NTCP_Overlap_Groups sheets. That workbook has been replaced
by data/Radiobiological_TCP_NTCP.xlsx, which unifies everything (n, m,
TD50, citation, verification status, notes) into one flat, per-record
sheet. This script has been rewritten end-to-end for that new schema; the
output JSON schema (core.lkb_bank.LKBBankEntry) is unchanged, so nothing
elsewhere in the app needed to change.

Run directly to rebuild the bank:
    python build_ntcp_bank_v2.py
"""
import json
import re
import sys
from pathlib import Path

import openpyxl

HERE = Path(__file__).resolve().parent
XLSX_PATH = HERE / "data" / "Radiobiological_TCP_NTCP.xlsx"
SHEET_NAME = "NTCP_Parameter_Bank"
OUT_PATH = HERE / "data" / "lkb_parameter_bank.json"

COLUMNS = [
    "record_id", "model_family", "year_period", "organ", "endpoint",
    "endpoint_timing", "volume_dose_metric", "n", "m", "td50",
    "td50_unit_ref", "alpha_beta", "cohort_modality", "sample_size",
    "evidence_type", "source_id", "authors", "title", "journal",
    "doi_link", "verification_status", "compat_notes",
]

# Verification Status (free text, new workbook) -> status key.
# All rows in this sheet are numerically complete (n, m, TD50 all present),
# so every status below is selectable=True; the key only changes the
# friendly label shown in "Add from LKB bank" (see
# ui/dialog_add_from_bank.py's status_note dict, kept in sync with these).
STATUS_MAP = {
    "Verified; unique parameter set in current bank": "verified_unique",
    "Verified; alternative/partial-overlap parameter set": "verified_alternative_overlap",
    "Verified; distinct endpoint (not a direct replacement)": "verified_distinct_endpoint",
    "Verified from source Table 2": "verified_table2",
    "Unverified - citation likely invalid, values plausible but "
    "unconfirmed (see notes)": "unverified_citation_caution",
}


def clean_organ_name(name: str) -> str:
    """Collapse redundant 'X (X)' style duplication (case-insensitive)."""
    name = (name or "").strip()
    m = re.match(r"^(.*?)\s*\((.*?)\)\s*$", name)
    if m and m.group(1).strip().casefold() == m.group(2).strip().casefold():
        return m.group(1).strip()
    return name


def extract_year(year_period_text) -> str:
    """Pull the last 4-digit (19xx/20xx) year out of a 'text | YYYY' or
    plain-YYYY cell."""
    text = str(year_period_text or "")
    full = re.findall(r"(?:19|20)\d{2}", text)
    return full[-1] if full else ""


def first_author_label(authors_text: str, year: str) -> str:
    authors_text = (authors_text or "").strip()
    if not authors_text:
        return year or ""
    first = authors_text.split(",")[0].strip()
    first = re.sub(r"\s+et\s+al\.?$", "", first, flags=re.IGNORECASE).strip()
    return f"{first} et al. {year}".strip() if year else f"{first} et al."


def parse_numeric(text):
    """Parse a numeric bank cell that may carry an annotation, e.g.
    '1.00 (constant)', '1 (fixed)', '0.12 (0.09-0.30)', '\u22481 (1.03\u00b10.17)'.
    Takes the FIRST numeric token in the cell (the point-estimate value
    the source itself leads with), ignoring surrounding symbols/units.
    Returns None for genuinely missing/non-numeric values (e.g. 'Not
    reported')."""
    if text is None:
        return None
    if isinstance(text, (int, float)):
        return float(text)
    text = str(text).strip()
    if not text or text.casefold().startswith("not reported"):
        return None
    m = re.search(r"[-+]?\d*\.?\d+", text)
    return float(m.group(0)) if m else None


def build_citation(authors, title, journal) -> str:
    parts = []
    authors = (str(authors).strip() if authors else "")
    title = (str(title).strip() if title else "")
    journal = (str(journal).strip() if journal and str(journal).casefold() != "nan" else "")
    if authors:
        parts.append(authors.rstrip(".") + ".")
    if title:
        parts.append(title.rstrip(".") + ".")
    if journal:
        parts.append(journal.rstrip(".") + ".")
    return " ".join(parts).strip()


def build_notes(compat_notes, evidence_type, verification_status, sample_size) -> str:
    parts = []
    if compat_notes:
        parts.append(str(compat_notes).strip())
    if evidence_type and str(evidence_type).strip():
        parts.append(f"Evidence type: {str(evidence_type).strip()}")
    if verification_status and str(verification_status).strip():
        parts.append(f"Verification: {str(verification_status).strip()}")
    if sample_size is not None and str(sample_size).strip() and str(sample_size).casefold() != "nan":
        parts.append(f"Sample size: {str(sample_size).strip()}")
    return " | ".join(parts)


def main():
    if not XLSX_PATH.exists():
        sys.exit(f"Evidence spreadsheet not found: {XLSX_PATH}")

    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        sys.exit(f"Sheet '{SHEET_NAME}' not found in {XLSX_PATH}. Sheets present: {wb.sheetnames}")
    ws = wb[SHEET_NAME]

    rows = list(ws.iter_rows(values_only=True))
    data_rows = rows[1:]

    records = []
    skipped = []
    seen_labels_per_group = {}

    for raw in data_rows:
        if raw is None or all(v is None for v in raw):
            continue
        rec = dict(zip(COLUMNS, raw))

        organ = clean_organ_name(rec["organ"])
        endpoint = str(rec["endpoint"] or "").strip()
        if rec["endpoint_timing"] and str(rec["endpoint_timing"]).strip():
            endpoint = f"{endpoint} ({str(rec['endpoint_timing']).strip()})"

        n_val = parse_numeric(rec["n"])
        m_val = parse_numeric(rec["m"])
        td50_val = parse_numeric(rec["td50"])
        if n_val is None or m_val is None or td50_val is None:
            status, selectable = "incomplete_reference_only", False
        else:
            status = STATUS_MAP.get(str(rec["verification_status"] or "").strip(), "unverified_citation_caution")
            selectable = True

        year = extract_year(rec["year_period"])
        label = first_author_label(rec["authors"], year)

        # Disambiguate labels that collide within the same organ+endpoint
        # group (the "Add from LKB bank" author dropdown keys off label
        # text alone within a group -- a collision would hide an entry).
        group_key = (organ, endpoint)
        seen = seen_labels_per_group.setdefault(group_key, set())
        base_label = label
        while label in seen:
            label = f"{base_label} [{rec['record_id']}]"
        seen.add(label)

        entry = {
            "organ": organ,
            "endpoint": endpoint,
            "label": label,
            "n": float(n_val) if n_val is not None else 0.0,
            "m": float(m_val) if m_val is not None else 0.0,
            "td50_gy": float(td50_val) if td50_val is not None else 0.0,
            "alpha_beta": parse_numeric(rec["alpha_beta"]),
            "vref": str(rec["volume_dose_metric"] or "").strip(),
            "dose_reference": str(rec["td50_unit_ref"] or "").strip(),
            "status": status,
            "selectable": selectable,
            "source": build_citation(rec["authors"], rec["title"], rec["journal"]),
            "url": (str(rec["doi_link"]).strip() if rec["doi_link"] and str(rec["doi_link"]).casefold() != "nan" else None),
            "notes": build_notes(rec["compat_notes"], rec["evidence_type"], rec["verification_status"], rec["sample_size"]),
            "model_id": str(rec["record_id"] or "").strip(),
            "cohort_modality": str(rec["cohort_modality"] or "").strip(),
        }

        if not selectable:
            skipped.append(entry["model_id"])
        records.append(entry)

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(OUT_PATH, "w", encoding="utf-8") as f:
        json.dump(records, f, indent=2, ensure_ascii=False)

    print(f"Wrote {len(records)} LKB/NTCP bank records to {OUT_PATH}")
    print(f"  selectable: {sum(1 for r in records if r['selectable'])}, "
          f"not selectable: {sum(1 for r in records if not r['selectable'])}")
    if skipped:
        print(f"  not selectable (incomplete n/m/TD50): {', '.join(skipped)}")


if __name__ == "__main__":
    main()
